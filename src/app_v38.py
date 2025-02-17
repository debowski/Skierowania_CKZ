import json
import os
# import subprocess
import sys
import threading
import pythoncom
import tkinter as tk
# import concurrent.futures

# import tkinter.ttk as ttk
from tkinter import StringVar, filedialog
import openpyxl
import ttkbootstrap as ttkb
# from docx import Document
from docx2pdf import convert
from docxtpl import DocxTemplate
# from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")


def convert_with_com(file_path):
    """Wrapper function to initialize COM for the conversion."""
    pythoncom.CoInitialize()  # Initialize COM for this thread
    try:
        convert(file_path)  # Call the conversion
    finally:
        pythoncom.CoUninitialize()  # Uninitialize COM


class App:

    def __init__(self):

        # Określenie ścieżki do folderu aplikacji
        self.app_dir = self.get_resource_path("")

        # Określenie ścieżki do folderu "Data"
        self.data_dir = os.path.join(self.app_dir, "Data")

        # Określenie ścieżki do folderu "Skierowania"
        self.skierowania_dir = os.path.join(self.app_dir, "Skierowania")

        # Określenie Ścieżki do pliku JSON
        self.json_file_path = os.path.join(self.app_dir, "zawody.json")

        # Załaduj dane z pliku JSON
        with open(self.json_file_path, "r", encoding="utf-8") as file:
            self.zawody_dict = json.load(file)

        self.root = ttkb.Window(themename="solar")
        self.root.title("Skierowania 0.38")
        self.root.grid()
        self.root.columnconfigure(0, weight=0, minsize=500)
        self.root.columnconfigure(1, weight=1, minsize=400)
        self.root.rowconfigure(0, weight=1)
        self.dodaj_widzety()
        self.credits()

    def get_resource_path(self, relative_path):
        """Zwraca ścieżkę do pliku, działającą w przypadku uruchomienia zarówno z .py jak i .exe"""
        try:
            base_path = sys._MEIPASS
        except AttributeError:
            # Jeżeli uruchamiany jest plik .py

            base_path = os.path.dirname(os.path.abspath(__file__))

        return os.path.join(base_path, relative_path)

    def load_zawody(self, file_path):
        try:
            with open(file_path, "r", encoding="utf-8") as file:
                return json.load(file)
        except FileNotFoundError:
            print(f"Błąd: Plik {file_path} nie został znaleziony.")
            return {}
        except json.JSONDecodeError:
            print(f"Błąd: Nieprawidłowy format pliku {file_path}.")
            return {}

    def wczytaj_indeksy_kolumn(self):
        wb = openpyxl.load_workbook(self.plik)
        sheet = wb.active

        headers = [cell.value for cell in sheet[1]]

        # Znajdź indeksy kolumn
        self.imie_idx = headers.index("Imię")
        self.drugie_imie_idx = headers.index("Drugie imię")
        self.nazwisko_idx = headers.index("Nazwisko")
        self.pesel_idx = headers.index("PESEL")
        self.data_urodzenia_idx = headers.index("Data urodzenia")
        self.oddzial_idx = headers.index("Dane oddziału")
        self.zawod_idx = headers.index("Specjalność/Zawód")

    def dodaj_widzety(self):
        self.frame = ttkb.Frame(self.root)
        self.frame.grid(row=0, column=0, sticky="nsew", padx=5, pady=5)

        self.frame2 = ttkb.Frame(self.root, bootstyle="success")
        self.frame2.grid(row=0, column=1, sticky="nsew", padx=10, pady=10)

        self.frame.columnconfigure(0, weight=1)
        self.frame.columnconfigure(1, weight=1)
        self.frame.rowconfigure(0, weight=0)
        self.frame.rowconfigure(9, weight=1)

        self.frame2.columnconfigure(0, weight=1)
        self.frame2.rowconfigure(0, weight=1)

        self.plik = ""

        self.btn_wyb_plik = ttkb.Button(
            self.frame,
            text="Wybierz plik",
            bootstyle="warning",
            command=self.otwarcie_pliku,
        )
        self.btn_wyb_plik.grid(
            row=0, column=0, sticky="nsew", columnspan=3, padx=5, pady=5
        )

        # Wybieranie klasy

        self.var = StringVar()

        self.radio1 = ttkb.Radiobutton(
            self.frame,
            text="Klasa 1",
            variable=self.var,
            value="1",
            bootstyle="success-outline-toolbutton",
            command=self.set_lista_zawodow_1,
        )
        self.radio2 = ttkb.Radiobutton(
            self.frame,
            text="Klasa 2",
            variable=self.var,
            value="2",
            bootstyle="success-outline-toolbutton",
            command=self.set_lista_zawodow_2,
        )
        self.radio3 = ttkb.Radiobutton(
            self.frame,
            text="Klasa 3",
            variable=self.var,
            value="3",
            bootstyle="success-outline-toolbutton",
            command=self.set_lista_zawodow_3,
        )
        self.radio1.grid(row=1, column=0, sticky="nsew", padx=5, pady=5)
        self.radio2.grid(row=1, column=1, sticky="nsew", padx=5, pady=5)
        self.radio3.grid(row=1, column=2, sticky="nsew", padx=5, pady=5)

        self.radiobuttons = []
        self.lista_zawodow = StringVar()

        self.zawody = ("Sprzedawca",)
        self.zawody1 = ("Wybierz zawód...",)
        self.zawody2 = ("Wybierz zawód...",)
        self.zawody3 = ("Wybierz zawód...",)

        # Set the default value for the combobox to be the first item in the list of values.
        self.combo_current_var = tk.StringVar()
        self.combobox = ttkb.Combobox(
            self.frame,
            values=self.zawody,
            textvariable=self.combo_current_var,
            bootstyle="success",
        )
        self.combobox.grid(row=2, column=0, sticky="nsew", padx=5, pady=5, columnspan=3)
        self.combobox.configure(state="readonly")
        self.combobox.set(self.zawody[0])
        self.combobox.current(0)

        self.combobox.bind("<<ComboboxSelected>>", self.wypisanie_osob)

        self.lab_data_wystawienia = ttkb.Label(self.frame, text="Data wystawienia")
        self.lab_data_wystawienia.grid(row=3, column=0, sticky="nsew", padx=5, pady=5)

        self.data_wystawienia = ttkb.DateEntry(self.frame, firstweekday=0, bootstyle="success")
        self.data_wystawienia.grid(
            row=3, column=1, sticky="nsew", padx=5, pady=5, columnspan=2
        )

        self.lab_data_rozpoczecia = ttkb.Label(self.frame, text="Data rozpoczęcia")
        self.lab_data_rozpoczecia.grid(row=4, column=0, sticky="nsew", padx=5, pady=5)

        self.data_rozpoczecia = ttkb.DateEntry(self.frame, firstweekday=0, bootstyle="success")
        self.data_rozpoczecia.grid(
            row=4, column=1, sticky="nsew", padx=5, pady=5, columnspan=2
        )

        self.lab_data_zakonczenia = ttkb.Label(self.frame, text="Data zakończenia")
        self.lab_data_zakonczenia.grid(row=5, column=0, sticky="nsew", padx=5, pady=5)

        self.data_zakonczenia = ttkb.DateEntry(self.frame, firstweekday=0, bootstyle="success")
        self.data_zakonczenia.grid(
            row=5, column=1, sticky="nsew", padx=5, pady=5, columnspan=2
        )

        self.label_godzina_rozpoczecia = ttkb.Label(
            self.frame, text="Godzina rozpoczęcia"
        )
        self.label_godzina_rozpoczecia.grid(
            row=6, column=0, sticky="nsew", padx=5, pady=5
        )

        self.godzina_rozpoczecia = ttkb.Spinbox(
            self.frame, from_=0, to=23, justify="center", format="%02.0f"
        )
        self.godzina_rozpoczecia.grid(row=6, column=1, sticky="nsew", padx=5, pady=5)
        self.godzina_rozpoczecia.insert(0, "08")

        self.minuty_rozpoczecia = ttkb.Spinbox(
            self.frame, from_=0, to=59, justify="center", format="%02.0f"
        )
        self.minuty_rozpoczecia.grid(row=6, column=2, sticky="nsew", padx=5, pady=5)
        self.minuty_rozpoczecia.insert(0, "00")

        self.btn_utworz_wykaz = ttkb.Button(
            self.frame,
            text="Utwórz wykaz",
            bootstyle="success",
            command=self.utworz_wykaz,
        )
        self.btn_utworz_wykaz.grid(row=7, column=0, sticky="nsew", padx=5, pady=5)

        self.btn_utworz_wykaz_pdf = ttkb.Button(
            self.frame,
            text="Konwersja do PDF",
            bootstyle="dark",
            command=self.utworz_wykaz_pdf,
        )
        self.btn_utworz_wykaz_pdf.grid(row=7, column=1, sticky="nsew", padx=5, pady=5)

        self.btn_otworz_folder_wykaz = ttkb.Button(
            self.frame, text="Otwórz folder wykazy", command=self.otworz_folder_wykazy, bootstyle="success"
        )
        self.btn_otworz_folder_wykaz.grid(
            row=7, column=2, sticky="nsew", padx=5, pady=5, columnspan=1
        )

        self.btn_utworz_skierowania = ttkb.Button(
            self.frame,
            text="Utwórz skierowania",
            bootstyle="success",
            command=self.utworz_skierowania,
        )
        self.btn_utworz_skierowania.grid(row=8, column=0, sticky="nsew", padx=5, pady=5)

        self.btn_utworz_skierowania_pdf = ttkb.Button(
            self.frame,
            text="Konwersja do PDF",
            bootstyle="dark",
            command=self.utworz_skierowania_pdf,
        )
        self.btn_utworz_skierowania_pdf.grid(
            row=8, column=1, sticky="nsew", padx=5, pady=5
        )

        self.btn_otworz_folder_skierowania = ttkb.Button(
            self.frame,
            text="Otwórz folder skierowania",
            command=self.otworz_folder_skierowania,
            bootstyle="success"
        )
        self.btn_otworz_folder_skierowania.grid(
            row=8, column=2, sticky="nsew", padx=5, pady=5, columnspan=1
        )

        self.wynik = ttkb.Label(self.frame, text="Wynik", bootstyle="inverse-dark")
        self.wynik.grid(row=9, column=0, sticky="sew", padx=5, pady=5, columnspan=3)

        # self.separator = ttkb.Separator(self.frame, orient="horizontal", bootstyle="success")
        # self.separator.grid(row=9, column=0, sticky="sew", padx=5, pady=5, columnspan=3)

        # frame2 - przyciski
        self.pole_tekstowe = tk.Text(self.frame2)
        self.pole_tekstowe.grid(row=0, column=0, padx=1, pady=1, sticky="nsew")

    def set_lista_zawodow_1(self) -> None:
        self.combobox["values"] = tuple(self.zawody1)
        self.wypisanie_osob()

    def set_lista_zawodow_2(self) -> None:
        self.combobox["values"] = tuple(self.zawody2)
        self.wypisanie_osob()

    def set_lista_zawodow_3(self) -> None:
        self.combobox["values"] = tuple(self.zawody3)
        self.wypisanie_osob()

    def utworz_lz(self):

        wb = openpyxl.load_workbook(self.plik)
        sheet = wb.active

        self.zawody1 = set()
        self.zawody2 = set()
        self.zawody3 = set()

        self.wczytaj_indeksy_kolumn()  # Ustawia indeksy kolumn na początku

        # Odczytaj pierwszy wiersz (nagłówki)
        headers = [cell.value for cell in sheet[1]]

        # Znajdź indeksy dla wszystkich potrzebnych kolumn
        try:
            #imie_idx = headers.index("Imię")
            #drugie_imie_idx = headers.index("Drugie imię")
            #nazwisko_idx = headers.index("Nazwisko")
            #pesel_idx = headers.index("PESEL")
            #data_urodzenia_idx = headers.index("Data urodzenia")
            oddzial_idx = headers.index("Dane oddziału")
            zawod_idx = headers.index("Specjalność/Zawód")
        except ValueError as e:
            raise ValueError(f"Nie znaleziono wymaganej kolumny: {e}")

        # Iteruj przez wiersze (pomijając nagłówek)
        for row in sheet.iter_rows(
            min_row=2
        ):  # Pomija nagłówki, zaczynając od drugiego wiersza
            # Pobieranie danych z kolumn
            # imie = row[imie_idx].value
            # drugie_imie = row[drugie_imie_idx].value
            # nazwisko = row[nazwisko_idx].value
            # pesel = row[pesel_idx].value
            # data_urodzenia = row[data_urodzenia_idx].value
            klasa = row[oddzial_idx].value.split()[0][
                0
            ]  # Pobiera pierwszy znak z kolumny "Dane oddziału"
            zawod = row[
                zawod_idx
            ].value  # Pobiera wartość z kolumny "Specjalność/Zawód"

            if klasa == "1":
                self.zawody1.add(zawod)
            elif klasa == "2":
                self.zawody2.add(zawod)
            elif klasa == "3":
                self.zawody3.add(zawod)

    def check_columns(self, file_path):
        expected_columns = [
            "Imię",
            "Drugie imię",
            "Nazwisko",
            "Dane oddziału",
            "PESEL",
            "Specjalność/Zawód",
        ]

        try:
            wb = openpyxl.load_workbook(file_path)
            sheet = wb.active

            # Pobierz pierwszy wiersz
            first_row = sheet.iter_rows(min_row=1, max_row=1)

            # Utwórz listę nazw kolumn
            columns = []

            for cell_tuple in first_row:
                for cell in cell_tuple:
                    if cell.value is not None:
                        columns.append(cell.value)

            if set(expected_columns).issubset(set(columns)):
                print("Plik ma poprawną strukturę kolumn.")
                self.btn_wyb_plik.configure(bootstyle="success")
                self.btn_utworz_wykaz.configure(state="normal")
                self.btn_utworz_skierowania.configure(state="normal")
                self.wczytaj_indeksy_kolumn()  # Ustawia indeksy kolumn na początku

                return True
            else:
                self.btn_wyb_plik.configure(text="Niepoprawne dane")
                self.btn_wyb_plik.configure(bootstyle="danger")

                self.btn_utworz_wykaz.configure(state="disabled")
                self.btn_utworz_skierowania.configure(state="disabled")

                print("Plik nie zawiera wszystkich oczekiwanych kolumn.")

                return False

        except Exception as e:
            print("Błąd podczas sprawdzania pliku:", e)
            return False

    def otwarcie_pliku(self):
        filetypes = (("Arkusze", "*.xlsx"), ("All files", "*.*"))

        self.plik = filedialog.askopenfilename(
            title="Wybierz plik", initialdir=self.data_dir, filetypes=filetypes
        )

        if self.plik and self.check_columns(self.plik):
            self.btn_wyb_plik.configure(text=self.plik)
            self.utworz_lz()
            self.radio1.invoke()
        else:
            print("Nieprawidłowa struktura kolumn w pliku lub nie wybrano pliku.")

    def filtruj_dane(self, plik, var, combobox):
        wb = openpyxl.load_workbook(open(plik, "rb"), read_only=True)
        sheet = wb.active


        # ======================================================================

        # # Odczytaj pierwszy wiersz (nagłówki)
        # headers = [cell.value for cell in sheet[1]]

        # # Znajdź indeksy dla wszystkich potrzebnych kolumn
        # try:
        #     imie_idx = headers.index("Imię")
        #     drugie_imie_idx = headers.index("Drugie imię")
        #     nazwisko_idx = headers.index("Nazwisko")
        #     pesel_idx = headers.index("PESEL")
        #     data_urodzenia_idx = headers.index("Data urodzenia")
        #     oddzial_idx = headers.index("Dane oddziału")
        #     zawod_idx = headers.index("Specjalność/Zawód")
        # except ValueError as e:
        #     raise ValueError(f"Nie znaleziono wymaganej kolumny: {e}")


        # ======================================================================


        # # Iteruj przez wiersze (pomijając nagłówek)
        # for row in sheet.iter_rows(
        #     min_row=2
        # ):  # Pomija nagłówki, zaczynając od drugiego wiersza
        #     # Pobieranie danych z kolumn
        #     imie = row[imie_idx].value
        #     drugie_imie = row[drugie_imie_idx].value
        #     nazwisko = row[nazwisko_idx].value
        #     pesel = row[pesel_idx].value
        #     data_urodzenia = row[data_urodzenia_idx].value
        #     klasa = row[oddzial_idx].value.split()[0][
        #         0
        #     ]  # Pobiera pierwszy znak z kolumny "Dane oddziału"
        #     zawod = row[
        #         zawod_idx
        #     ].value  # Pobiera wartość z kolumny "Specjalność/Zawód"

        filtered_data = []
        for row in sheet.iter_rows(min_row=2):  # Dodano min_row=2, aby pominąć nagłówki
            if (
                row[self.oddzial_idx].value
                and var.lower() in row[self.oddzial_idx].value.lower()
                and row[self.zawod_idx].value
                and combobox.lower() in row[self.zawod_idx].value.lower()
            ):
                filtered_data.append(row)

        return filtered_data
    def wyswietl_dane(self, filtered_data):
        # Create a list comprehension to quickly generate the text
        
        tekst = "\n".join(
            [
                f"{i+1}. {row[self.imie_idx].value} {row[self.drugie_imie_idx].value} {row[self.nazwisko_idx].value}"
                for i, row in enumerate(filtered_data)
            ]
        )

        # Insert the text into the text widget
        self.pole_tekstowe.delete(1.0, tk.END)
        
        if tekst == "":
            self.pole_tekstowe.insert(tk.END, "Brak danych")
        else:
            self.pole_tekstowe.insert(tk.END, tekst)

    def wypisanie_osob(self, event=None):
        if self.plik == "":
            # Wywołaj funkcję obsługi braku pliku
            self.brak_pliku()
            return

        filtered_data = self.filtruj_dane(
            self.plik, self.var.get(), self.combobox.get()
        )
        self.wyswietl_dane(filtered_data)

    def brak_pliku(self):
        self.pole_tekstowe.delete(1.0, tk.END)
        self.pole_tekstowe.insert(tk.END, "Nie wybrano pliku")

    def symbolZawodu(self, specjalnosc) -> str:
        # lista zawodów została przeniesiona do pliku zawody.json
        return self.zawody_dict.get(specjalnosc, "N/A")

    def znajdz_indeksy_kolumn(self, arkusz, naglowki):
        # ta funkcja służy do pobrania indeksoów kolumn z arkusza Excela na podstawie nagłówków

        indeksy_kolumn = {}
        for wiersz in range(1, arkusz.max_row + 1):
            for kolumna in range(1, arkusz.max_column + 1):
                komorka = arkusz.cell(row=wiersz, column=kolumna)
                if komorka.value in naglowki:
                    indeksy_kolumn[komorka.value] = kolumna
        return indeksy_kolumn

    def utworz_wykaz(self):
        # # cwd = os.path.dirname(__file__)
        # parent_dir = os.path.dirname(cwd)
        # # folder_path = os.path.abspath(
        # #     os.path.join(os.path.dirname(__file__), parent_dir, "Data", "Wykazy")
        # # )

        # folder_path = os.path.join(self.app_dir, "Data" ,"Wykazy")

        wybrane_dane = self.filtruj_dane(self.plik, self.var.get(), self.combobox.get())

        doc = DocxTemplate(
            os.path.join(self.app_dir, "Szablony", "szablon_wykaz_v3.docx")
        )

        # naglowki = ['Imię', 'Nazwisko', 'PESEL', 'Data urodzenia', 'Miejsce urodzenia']
        naglowki = ["Imię", "Nazwisko", "PESEL"]

        wb = openpyxl.load_workbook(open(self.plik, "rb"))
        sheet = wb.active
        indeksy_kolumn = self.znajdz_indeksy_kolumn(sheet, naglowki)

        lista = ""

        for linia in range(len(wybrane_dane)):
            rekord = wybrane_dane[linia]

            lista = (
                lista
                + str(linia + 1)
                + ". "
                + rekord[indeksy_kolumn["Imię"] - 1].value
                + " "
                + rekord[indeksy_kolumn["Nazwisko"] - 1].value
                + "\n"
            )

        # Tworzenie kontekstu dla dalszego użycia (np. w szablonach)
        context = {
            "dataWyst": self.data_wystawienia.entry.get(),
            "imie": (
                rekord[indeksy_kolumn["Imię"] - 1].value
                if "Imię" in indeksy_kolumn
                else ""
            ),
            "drugie_imie": (
                rekord[indeksy_kolumn["Drugie imię"] - 1].value
                if "Drugie imię" in indeksy_kolumn
                else ""
            ),
            "nazwisko": (
                rekord[indeksy_kolumn["Nazwisko"] - 1].value
                if "Nazwisko" in indeksy_kolumn
                else ""
            ),
            "PESEL": (
                rekord[indeksy_kolumn["PESEL"] - 1].value
                if "PESEL" in indeksy_kolumn
                else ""
            ),
            "zawod": self.combobox.get(),
            "kodZawodu": self.symbolZawodu(self.combobox.get()),
            "dataRozp": self.data_rozpoczecia.entry.get(),
            "dataZako": self.data_zakonczenia.entry.get(),
            "godzRozp": f"{self.godzina_rozpoczecia.get()}:{self.minuty_rozpoczecia.get()}",
            "stopien": self.var.get(),
            "tabela": lista,
        }

        doc.render(context)

        if not os.path.exists(os.path.join(self.app_dir, "Data")):
            os.mkdir(os.path.join(self.data_dir, "Data"))
        if not os.path.exists(os.path.join(self.app_dir, "Data", "Wykazy")):
            os.mkdir(os.path.join(self.app_dir, "Data", "Wykazy"))

        doc.save(
            os.path.join(
                self.app_dir,
                "Data",
                "Wykazy",
                f"{context['stopien']}_{context['zawod']}.docx",
            )
        )

        self.wynik.configure(
            text=f"Utworzono wykaz zawierający: {str(linia + 1)} pozycji"
        )

        # Ustawienie napisu na przycisku do generowania pdf
        folder_path = os.path.abspath(os.path.join(self.app_dir, "Data", "Wykazy"))

        files_to_convert = os.listdir(folder_path)
        total_files_wykazy = sum(
            1
            for file_name in files_to_convert
            if os.path.isfile(os.path.join(folder_path, file_name))
        )

        self.btn_utworz_wykaz_pdf.configure(
            text=f"PDF: {str(total_files_wykazy)} plików"
        )

    def utworz_skierowania(self):
        wybrane_dane = self.filtruj_dane(self.plik, self.var.get(), self.combobox.get())
        doc = DocxTemplate(
            os.path.join(self.app_dir, "Szablony", "szablon_skierowania_v3.docx")
        )

        # naglowki = ['Imię', 'Nazwisko', 'PESEL', 'Data urodzenia', 'Miejsce urodzenia']
        naglowki = ["Imię", "Drugie imię", "Nazwisko", "PESEL"]
        wb = openpyxl.load_workbook(open(self.plik, "rb"))
        sheet = wb.active
        indeksy_kolumn = self.znajdz_indeksy_kolumn(sheet, naglowki)

        lista = ""

        for linia in range(len(wybrane_dane)):
            rekord = wybrane_dane[linia]

            context = {
                "dataWyst": self.data_wystawienia.entry.get(),
                "imie": rekord[indeksy_kolumn["Imię"] - 1].value,
                "drugie_imie": rekord[indeksy_kolumn["Drugie imię"] - 1].value,
                "nazwisko": rekord[indeksy_kolumn["Nazwisko"] - 1].value,
                # 'dataUrodzenia': rekord[indeksy_kolumn['Data urodzenia']-1].value,
                # 'miejsceUrodzenia': rekord[indeksy_kolumn['Miejsce urodzenia']-1].value,
                "PESEL": rekord[indeksy_kolumn["PESEL"] - 1].value,
                "zawod": self.combobox.get(),
                "kodZawodu": self.symbolZawodu(self.combobox.get()),
                "dataRozp": self.data_rozpoczecia.entry.get(),
                "dataZako": self.data_zakonczenia.entry.get(),
                "godzRozp": self.godzina_rozpoczecia.get()
                + ":"
                + self.minuty_rozpoczecia.get(),
                "stopien": self.var.get(),
                "tabela": lista,
            }

            doc.render(context)

            if not os.path.exists(os.path.join(self.app_dir, "Data")):
                os.mkdir(os.path.join(self.app_dir, "Data"))

            if not os.path.exists(os.path.join(self.app_dir, "Data", "Skierowania")):
                os.mkdir(os.path.join(self.app_dir, "Data", "Skierowania"))

            nazwa_pliku = f"{context['stopien']}_{context['zawod']}{context['imie']}{context['nazwisko']}.docx"
            sciezka_pliku = os.path.join(
                self.app_dir, "Data", "Skierowania", nazwa_pliku
            )

            # Sprawdzamy czy plik już istnieje
            if os.path.exists(sciezka_pliku):
                # Jeśli plik istnieje, dodajemy numer do nazwy pliku
                numer = 1
                while os.path.exists(sciezka_pliku):
                    nazwa_pliku = f"{context['stopien']}_{context['zawod']}{context['imie']}{context['nazwisko']}_{numer}.docx"
                    sciezka_pliku = os.path.join(
                        self.app_dir, "Data", "Skierowania", nazwa_pliku
                    )
                    numer += 1

            doc.save(sciezka_pliku)

            wb.close()

        # informacja zwrotna
        self.wynik.configure(text=f"Utworzono: {str(linia + 1)} dokumentów")
        # folder_path_skierowania = os.path.join(self.app_dir, "Data", "Skierowania")

        total_files_skierowania = self.zlicz_plik_docx(
            os.path.join(self.app_dir, "Data", "Skierowania")
        )
        print(f"Total files: {total_files_skierowania}")

        self.btn_utworz_skierowania_pdf.configure(
            text=f"PDF: {str(total_files_skierowania)} plików"
        )

        # Ustawienie napisu na przycisku do generowania pdf
        total_files_skierowania = self.zlicz_plik_docx(
            os.path.join(self.app_dir, "Data", "Skierowania")
        )

        self.btn_utworz_skierowania_pdf.configure(
            text=f"PDF: {str(total_files_skierowania)} plików"
        )

    def zlicz_plik_docx(self, folder):
        """Zlicza pliki .docx w podanym folderze."""
        licznik = 0
        for plik in os.listdir(folder):
            if plik.endswith(".docx") and not plik.startswith("~"):
                licznik += 1
        return licznik

    def utworz_wykaz_pdf(self):
        """Generates PDF documents based on Word documents in the 'Wykazy' folder."""
        folder_path = os.path.join(self.app_dir, "Data", "Wykazy")

        # List of files to convert, ignoring temporary files
        pliki_do_konwersji = [
            plik
            for plik in os.listdir(folder_path)
            if plik.endswith(".docx") and not plik.startswith("~")
        ]

        # Convert only if there are files to convert
        if pliki_do_konwersji:
            for plik in pliki_do_konwersji:
                sciezka_do_pliku = os.path.join(folder_path, plik)
                thread = threading.Thread(
                    target=convert_with_com, args=(sciezka_do_pliku,)
                )
                thread.start()  # Start conversion in a separate thread
                print(f"Rozpoczęto konwersję pliku: {plik}")
        else:
            print("Brak plików do przetworzenia.")

    def utworz_skierowania_pdf(self):
        """Generates PDF documents based on Word documents in the 'Skierowania' folder."""
        folder_path = os.path.join(self.app_dir, "Data", "Skierowania")

        # List of files to convert, ignoring temporary files
        pliki_do_konwersji = [
            plik
            for plik in os.listdir(folder_path)
            if plik.endswith(".docx") and not plik.startswith("~")
        ]

        # Convert only if there are files to convert
        if pliki_do_konwersji:
            for plik in pliki_do_konwersji:
                sciezka_do_pliku = os.path.join(folder_path, plik)
                thread = threading.Thread(
                    target=convert_with_com, args=(sciezka_do_pliku,)
                )
                thread.start()  # Start conversion in a separate thread
                print(f"Rozpoczęto konwersję pliku: {plik}")
        else:
            print("Brak plików do przetworzenia.")

    def otworz_folder_wykazy(self):
        path = os.path.join(self.app_dir, "Data", "Wykazy")
        os.startfile(path)

    def otworz_folder_skierowania(self):
        path = os.path.join(self.app_dir, "Data", "Skierowania")
        os.startfile(path)

    def credits(self):
        """Wyświetlanie informacji o autorze i licencji."""
        version_number = self.root.title()
        credits_text = (
            "Autor: Piotr Dębowski\n"
            "Zespół Szkół Energetycznych i Usługowych w Łaziskch Górnych\n\n"
            f"Wersja: {version_number}\n\n"
        )

        license_text = "Ten program jest udostępniony na zasadach Licencji MIT. Oznacza to, że jesteś uprawniony do korzystania z niego w ramach warunków określonych w tej licencji.\n\n"
        disclaimer_text = "Autor tego programu nie ponosi odpowiedzialności za ewentualne szkody wynikające z jego użytkowania. Chociaż zrobiłem wszystko, co w mojej mocy, aby zapewnić poprawność i użyteczność tego oprogramowania, nie możgę zagwarantować jego bezbłędnego działania we wszystkich sytuacjach. Korzystając z tego programu, akceptujesz ryzyko związane z jego użyciem.\n\n"
        app_folders_tekst ="\n\nFoldery aplikacji:\n\n" + self.app_dir
        data_folders_tekst ="\n" + os.path.join(self.app_dir, "Data")
        skierowania_folders_tekst ="\n" + os.path.join(self.app_dir, "Data", "Skierowania")
        wykazy_folders_tekst ="\n" + os.path.join(self.app_dir, "Data", "Wykazy")
        credits_text += license_text + disclaimer_text + app_folders_tekst + data_folders_tekst + skierowania_folders_tekst + wykazy_folders_tekst
        self.pole_tekstowe.delete("1.0", tk.END)
        self.pole_tekstowe.insert(tk.END, credits_text)


if __name__ == "__main__":
    app = App()
    app.root.mainloop()
