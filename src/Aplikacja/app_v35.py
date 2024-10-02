import threading
import time
from docx2pdf import convert
from docxtpl import DocxTemplate
import os
import subprocess
import tkinter as tk
import tkinter.ttk as ttk
import ttkbootstrap as ttkb
from tkinter import StringVar
from tkinter import filedialog
import openpyxl
import sys
from docx import Document
import tkinter.simpledialog

sys.stdout.reconfigure(encoding='utf-8')
 

class App:

    def __init__(self):
        self.root = ttkb.Window(themename="darkly")
        self.root.title("Skierowania 0.35")
        self.root.grid()
        self.root.columnconfigure(0, weight=0, minsize=500)
        self.root.columnconfigure(1, weight=1, minsize=400)
        self.root.rowconfigure(0, weight=1)
        self.dodaj_widzety()
        self.credits()     
        
    def wczytaj_indeksy_kolumn(self):
        wb = openpyxl.load_workbook(self.plik)
        sheet = wb.active

        headers = [cell.value for cell in sheet[1]]

        # Znajdź indeksy kolumn
        self.imie_idx = headers.index('Imię')
        self.drugie_imie_idx = headers.index('Drugie imię')
        self.nazwisko_idx = headers.index('Nazwisko')
        self.pesel_idx = headers.index('PESEL')
        self.data_urodzenia_idx = headers.index('Data urodzenia')
        self.oddzial_idx = headers.index('Dane oddziału')
        self.zawod_idx = headers.index('Specjalność/Zawód')
        
        
        
        

    def dodaj_widzety(self):
        self.frame = ttkb.Frame(self.root)
        self.frame.grid(row=0, column=0, sticky="nsew", padx=5, pady=5)

        self.frame2 = ttkb.Frame(self.root, bootstyle="success")
        self.frame2.grid(row=0, column=1, sticky="nsew", padx=10, pady=10)

        self.frame.columnconfigure(0, weight=1)
        self.frame.columnconfigure(1, weight=1)
        self.frame.rowconfigure(0, weight=0)
        self.frame.rowconfigure(9, weight=1)

        self.frame2.columnconfigure(0, weight=1)
        self.frame2.rowconfigure(0, weight=1)

        self.plik = ""

        self.btn_wyb_plik = ttkb.Button(
            self.frame, text="Wybierz plik", bootstyle="info", command=self.otwarcie_pliku)
        self.btn_wyb_plik.grid(row=0, column=0, sticky="nsew",
                               columnspan=3, padx=5, pady=5)

        # Wybieranie klasy

        self.var = StringVar()

        self.radio1 = ttkb.Radiobutton(
            self.frame, text="Klasa 1", variable=self.var, value="1", bootstyle="info-outline-toolbutton", command=self.set_lista_zawodow_1)
        self.radio2 = ttkb.Radiobutton(
            self.frame, text="Klasa 2", variable=self.var, value="2", bootstyle="info-outline-toolbutton", command=self.set_lista_zawodow_2)
        self.radio3 = ttkb.Radiobutton(
            self.frame, text="Klasa 3", variable=self.var, value="3", bootstyle="info-outline-toolbutton", command=self.set_lista_zawodow_3)
        self.radio1.grid(row=1, column=0, sticky="nsew", padx=5, pady=5)
        self.radio2.grid(row=1, column=1, sticky="nsew", padx=5, pady=5)
        self.radio3.grid(row=1, column=2, sticky="nsew", padx=5, pady=5)

        self.radiobuttons = []
        self.lista_zawodow = StringVar()

        self.zawody = ('Sprzedawca',)
        self.zawody1 = ('Wybierz mądrze...',)
        self.zawody2 = ('Wybierz mądrze...',)
        self.zawody3 = ('Wybierz mądrze...',)

        # wstaw combobox z wyborem zawodow

        # Set the default value for the combobox to be the first item in the list of values.
        self.combo_current_var = tk.StringVar()
        self.combobox = ttkb.Combobox(
            self.frame, values=self.zawody, textvariable=self.combo_current_var, bootstyle="success")
        self.combobox.grid(row=2, column=0, sticky="nsew",
                           padx=5, pady=5, columnspan=3)
        self.combobox.configure(state='readonly')
        self.combobox.set(self.zawody[0])
        self.combobox.current(0)

        self.combobox.bind("<<ComboboxSelected>>", self.wypisanie_osob)

        self.lab_data_wystawienia = ttkb.Label(
            self.frame, text="Data wystawienia")
        self.lab_data_wystawienia.grid(
            row=3, column=0, sticky="nsew", padx=5, pady=5)

        self.data_wystawienia = ttkb.DateEntry(
            self.frame, firstweekday=0)
        self.data_wystawienia.grid(
            row=3, column=1, sticky="nsew", padx=5, pady=5, columnspan=2)

        self.lab_data_rozpoczecia = ttkb.Label(
            self.frame, text="Data rozpoczęcia")
        self.lab_data_rozpoczecia.grid(
            row=4, column=0, sticky="nsew", padx=5, pady=5)

        self.data_rozpoczecia = ttkb.DateEntry(
            self.frame)
        self.data_rozpoczecia.grid(
            row=4, column=1, sticky="nsew", padx=5, pady=5, columnspan=2)

        self.lab_data_zakonczenia = ttkb.Label(
            self.frame, text="Data zakończenia")
        self.lab_data_zakonczenia.grid(
            row=5, column=0, sticky="nsew", padx=5, pady=5)

        self.data_zakonczenia = ttkb.DateEntry(
            self.frame)
        self.data_zakonczenia.grid(
            row=5, column=1, sticky="nsew", padx=5, pady=5, columnspan=2)

        self.label_godzina_rozpoczecia = ttkb.Label(
            self.frame, text="Godzina rozpoczęcia")
        self.label_godzina_rozpoczecia.grid(
            row=6, column=0, sticky="nsew", padx=5, pady=5)

        self.godzina_rozpoczecia = ttkb.Spinbox(
            self.frame, from_=0, to=23, justify="center", format="%02.0f")
        self.godzina_rozpoczecia.grid(
            row=6, column=1, sticky="nsew", padx=5, pady=5)
        self.godzina_rozpoczecia.insert(0, "08")

        self.minuty_rozpoczecia = ttkb.Spinbox(
            self.frame, from_=0, to=59, justify="center", format="%02.0f")
        self.minuty_rozpoczecia.grid(
            row=6, column=2, sticky="nsew", padx=5, pady=5)
        self.minuty_rozpoczecia.insert(0, "00")

        self.btn_utworz_wykaz = ttkb.Button(
            self.frame, text="Utwórz wykaz", bootstyle="success", command=self.utworz_wykaz)
        self.btn_utworz_wykaz.grid(
            row=7, column=0, sticky="nsew", padx=5, pady=5)

        self.btn_utworz_wykaz_pdf = ttkb.Button(
            self.frame, text="Konwersja do PDF", bootstyle="info", command=self.utworz_wykaz_pdf)
        self.btn_utworz_wykaz_pdf.grid(
            row=7, column=1, sticky="nsew", padx=5, pady=5)

        self.btn_otworz_folder_wykaz = ttkb.Button(
            self.frame, text="Otwórz folder wykaz", command=self.otworz_folder_wykaz)
        self.btn_otworz_folder_wykaz.grid(
            row=7, column=2, sticky="nsew", padx=5, pady=5, columnspan=1)

        self.btn_utworz_skierowania = ttkb.Button(
            self.frame, text="Utwórz skierowania", bootstyle="success", command=self.utworz_skierowania)
        self.btn_utworz_skierowania.grid(
            row=8, column=0, sticky="nsew", padx=5, pady=5)

        self.btn_utworz_skierowania_pdf = ttkb.Button(
            self.frame, text="Konwersja do PDF", bootstyle="info", command=self.utworz_skierowania_pdf)
        self.btn_utworz_skierowania_pdf.grid(
            row=8, column=1, sticky="nsew", padx=5, pady=5)

        self.btn_otworz_folder_skierowania = ttkb.Button(
            self.frame, text="Otwórz folder skierowania", command=self.otworz_folder_skierowania)
        self.btn_otworz_folder_skierowania.grid(
            row=8, column=2, sticky="nsew", padx=5, pady=5, columnspan=1)

        self.wynik = ttkb.Label(self.frame, text="Wynik",
                                bootstyle="inverse-dark")
        self.wynik.grid(row=9, column=0, sticky="sew",
                        padx=5, pady=5, columnspan=3)

        # self.separator = ttkb.Separator(self.frame, orient="horizontal", bootstyle="success")
        # self.separator.grid(row=9, column=0, sticky="sew", padx=5, pady=5, columnspan=3)

        # frame2 - przyciski
        self.pole_tekstowe = tk.Text(self.frame2)
        self.pole_tekstowe.grid(
            row=0, column=0,  padx=1, pady=1, sticky="nsew")

    def set_lista_zawodow_1(self) -> None:
        self.combobox['values'] = list(self.zawody1)
        self.wypisanie_osob()

    def set_lista_zawodow_2(self) -> None:
        self.combobox['values'] = list(self.zawody2)
        self.wypisanie_osob()

    def set_lista_zawodow_3(self) -> None:
        self.combobox['values'] = list(self.zawody3)
        self.wypisanie_osob()

    def utworz_lz(self):

        wb = openpyxl.load_workbook(self.plik)
        sheet = wb.active

        self.zawody1 = set()
        self.zawody2 = set()
        self.zawody3 = set()
        
        self.wczytaj_indeksy_kolumn()  # Ustawia indeksy kolumn na początku
        
    # Odczytaj pierwszy wiersz (nagłówki)
        headers = [cell.value for cell in sheet[1]]

    # Znajdź indeksy dla wszystkich potrzebnych kolumn
        try:
            imie_idx = headers.index('Imię')
            drugie_imie_idx = headers.index('Drugie imię')
            nazwisko_idx = headers.index('Nazwisko')
            pesel_idx = headers.index('PESEL')
            data_urodzenia_idx = headers.index('Data urodzenia')
            oddzial_idx = headers.index('Dane oddziału')
            zawod_idx = headers.index('Specjalność/Zawód')
        except ValueError as e:
            raise ValueError(f"Nie znaleziono wymaganej kolumny: {e}")
        
        # Iteruj przez wiersze (pomijając nagłówek)
        for row in sheet.iter_rows(min_row=2):  # Pomija nagłówki, zaczynając od drugiego wiersza
            # Pobieranie danych z kolumn
            imie = row[imie_idx].value
            drugie_imie = row[drugie_imie_idx].value
            nazwisko = row[nazwisko_idx].value
            pesel = row[pesel_idx].value
            data_urodzenia = row[data_urodzenia_idx].value
            klasa = row[oddzial_idx].value.split()[0][0]  # Pobiera pierwszy znak z kolumny "Dane oddziału"
            zawod = row[zawod_idx].value  # Pobiera wartość z kolumny "Specjalność/Zawód"
           

            if klasa == "1":
                self.zawody1.add(zawod)
            elif klasa == "2":
                self.zawody2.add(zawod)
            elif klasa == "3":
                self.zawody3.add(zawod)

    # def check_columns(self, file_path):
    #     expected_columns = ['PESEL', 'Data urodzenia', 'Specjalność/Zawód',
    #                         'Miejsce urodzenia', 'Imię', 'Dane oddziału', 'Nazwisko']



    # po zmianie na rok szkolny 2024 / 2025 usunięto datę utodzenia i miejsce urodzenia
    def check_columns(self, file_path):
        expected_columns = ['PESEL', 'Specjalność/Zawód', 'Imię', 'Drugie imię', 'Dane oddziału', 'Nazwisko']

        try:
            wb = openpyxl.load_workbook(file_path)
            sheet = wb.active

            # Pobierz pierwszy wiersz
            first_row = sheet.iter_rows(min_row=1, max_row=1)

            # Utwórz listę nazw kolumn
            columns = []

            for cell_tuple in first_row:
                for cell in cell_tuple:
                    if cell.value is not None:
                        columns.append(cell.value)

            if set(expected_columns).issubset(set(columns)):
                print("Plik ma poprawną strukturę kolumn.")
                self.btn_wyb_plik.configure(bootstyle="success")
                self.btn_utworz_wykaz.configure(state="normal")
                self.btn_utworz_skierowania.configure(state="normal")
                self.wczytaj_indeksy_kolumn()  # Ustawia indeksy kolumn na początku

                return True
            else:
                self.btn_wyb_plik.configure(text="Niepoprawne dane")
                self.btn_wyb_plik.configure(bootstyle="danger")

                self.btn_utworz_wykaz.configure(state="disabled")
                self.btn_utworz_skierowania.configure(state="disabled")

                print("Plik nie zawiera wszystkich oczekiwanych kolumn.")

                return False

        except Exception as e:
            print("Błąd podczas sprawdzania pliku:", e)
            return False

    def otwarcie_pliku(self):
        filetypes = (
            ('Arkusze', '*.xlsx'),
            ('All files', '*.*')
        )

        print(os.getcwd())

        self.plik = filedialog.askopenfilename(title='Wybierz plik',
                                               initialdir='..\\Data',
                                               filetypes=filetypes)

        if self.plik:  # Sprawdzamy, czy plik został wybrany
            # Wywołujemy funkcję sprawdzającą kolumny
            if self.check_columns(self.plik):
                self.btn_wyb_plik.configure(text=self.plik)
                self.utworz_lz()
                self.radio1.invoke()
            else:
                # Jeśli funkcja check_columns zwróci False, to znaczy że struktura kolumn jest nieprawidłowa
                print("Nieprawidłowa struktura kolumn w pliku.")
        else:
            # Jeśli użytkownik nie wybrał pliku
            print("Nie wybrano pliku.")

    def filtruj_dane(self, plik, var, combobox):
        wb = openpyxl.load_workbook(open(plik, "rb"))
        sheet = wb.active

    # Odczytaj pierwszy wiersz (nagłówki)
        headers = [cell.value for cell in sheet[1]]

    # Znajdź indeksy dla wszystkich potrzebnych kolumn
        try:
            imie_idx = headers.index('Imię')
            drugie_imie_idx = headers.index('Drugie imię')
            nazwisko_idx = headers.index('Nazwisko')
            pesel_idx = headers.index('PESEL')
            data_urodzenia_idx = headers.index('Data urodzenia')
            oddzial_idx = headers.index('Dane oddziału')
            zawod_idx = headers.index('Specjalność/Zawód')
        except ValueError as e:
            raise ValueError(f"Nie znaleziono wymaganej kolumny: {e}")
        
        # Iteruj przez wiersze (pomijając nagłówek)
        for row in sheet.iter_rows(min_row=2):  # Pomija nagłówki, zaczynając od drugiego wiersza
            # Pobieranie danych z kolumn
            imie = row[imie_idx].value
            drugie_imie = row[drugie_imie_idx].value
            nazwisko = row[nazwisko_idx].value
            pesel = row[pesel_idx].value
            data_urodzenia = row[data_urodzenia_idx].value
            klasa = row[oddzial_idx].value.split()[0][0]  # Pobiera pierwszy znak z kolumny "Dane oddziału"
            zawod = row[zawod_idx].value  # Pobiera wartość z kolumny "Specjalność/Zawód"
 
 
        # filtered_data = []
        # for row in sheet.iter_rows():
        #     if (var.lower() in row[oddzial_idx].value.lower() and
        #             combobox.lower() in row[zawod_idx].value.lower()):
        #         filtered_data.append(row)
      
        filtered_data = []
        for row in sheet.iter_rows(min_row=2):  # Dodano min_row=2, aby pominąć nagłówki
            if (row[self.oddzial_idx].value and var.lower() in row[self.oddzial_idx].value.lower() and
                    row[self.zawod_idx].value and combobox.lower() in row[self.zawod_idx].value.lower()):
                filtered_data.append(row)      
        
        return filtered_data

    def wyswietl_dane(self, filtered_data):
        tekst = ""
        numer = 1
        for row in filtered_data:
            rekord = {
                "Imię": row[self.imie_idx].value,
                "Drugie imię": row[self.drugie_imie_idx].value,
                "Nazwisko": row[self.nazwisko_idx].value,
                "PESEL": row[self.pesel_idx].value,
                "Data urodzenia": row[self.data_urodzenia_idx].value,
                "Dane oddziału": row[self.oddzial_idx].value,
                "Specjalność/Zawód": row[self.zawod_idx].value
            }
            tekst += f"{numer}. {rekord['Imię']} {rekord['Drugie imię']} {rekord['Nazwisko']}\n"
            numer += 1

        # Wstawianie tekstu do ramki
        self.pole_tekstowe.delete(1.0, tk.END)
        self.pole_tekstowe.insert(tk.END, tekst)

    def wypisanie_osob(self, event=None):
        if self.plik == "":
            # Wywołaj funkcję obsługi braku pliku (zakładając, że jest zdefiniowana gdzie indziej)
            self.brak_pliku()
            return

        filtered_data = self.filtruj_dane(
            self.plik, self.var.get(), self.combobox.get())
        self.wyswietl_dane(filtered_data)

    def brak_pliku(self):
        self.pole_tekstowe.delete(1.0, tk.END)
        self.pole_tekstowe.insert(tk.END, "Nie wybrano pliku")

    def symbolZawodu(self, specjalnosc) -> str:
        

                
        zawody_dict = {
            "Administrator produkcji filmowej i telewizyjnej (Wprowadzono na podstawie rozporządzenia z dnia 18 stycznia 2023 r.)": "343919",
            "Animator rynku książki (Wprowadzono na podstawie rozporządzenia z dnia 18 stycznia 2023 r.)": "343305",
            "Asystent kierownika produkcji filmowej i telewizyjnej": "343902",
            "Asystent osoby niepełnosprawnej": "341201",
            "Asystentka stomatologiczna": "325101",
            "Automatyk": "731107",
            "Betoniarz-zbrojarz": "711402",
            "Blacharz": "721301",
            "Blacharz samochodowy": "721306",
            "Cieśla": "711501",
            "Cukiernik": "751201",
            "Dekarz": "712101",
            "Drukarz fleksograficzny": "732209",
            "Drukarz offsetowy": "732210",
            "Elektromechanik": "741201",
            "Elektromechanik pojazdów samochodowych": "741203",
            "Elektronik": "742117",
            "Elektryk": "741103",
            "Florysta": "343203",
            "Fotograf": "343101",
            "Fryzjer": "514101",
            "Garbarz skór": "753501",
            "Górnik eksploatacji otworowej": "811301",
            "Górnik eksploatacji podziemnej": "811101",
            "Górnik odkrywkowej eksploatacji złóż": "811102",
            "Górnik podziemnej eksploatacji kopalin innych niż węgiel kamienny": "811112",
            "Higienistka stomatologiczna": "325102",
            "Jeździec": "516408",
            "Kaletnik": "753702",
            "Kamieniarz": "711301",
            "Kelner": "513101",
            "Kierowca mechanik": "832201",
            "Kominiarz": "713303",
            "Koszykarz-plecionkarz": "731702",
            "Kowal": "722101",
            "Krawiec": "753105",
            "Kucharz": "512001",
            "Kuśnierz": "753106",
            "Lakiernik samochodowy": "713203",
            "Magazynier-logistyk": "432106",
            "Mechanik motocyklowy": "723107",
            "Mechanik pojazdów kolejowych (Opracowano na podstawie dokumentu z dnia 9 kwietnia 2020 r.)": "723318",
            "Mechanik pojazdów samochodowych": "723103",
            "Mechanik precyzyjny": "731103",
            "Mechanik-monter maszyn i urządzeń": "723310",
            "Mechanik-operator maszyn do produkcji drzewnej": "817212",
            "Mechanik-operator pojazdów i maszyn rolniczych": "834103",
            "Modelarz odlewniczy": "721104",
            "Monter budownictwa wodnego": "711701",
            "Monter izolacji budowlanych": "712401",
            "Monter izolacji przemysłowych": "712403",
            "Monter jachtów i łodzi": "711505",
            "Monter kadłubów jednostek pływających": "721406",
            "Monter konstrukcji budowlanych": "711102",
            "Monter nawierzchni kolejowej": "711603",
            "Monter sieci i instalacji sanitarnych": "712618",
            "Monter sieci i urządzeń telekomunikacyjnych": "742202",
            "Monter stolarki budowlanej": "712906",
            "Monter systemów rurociągowych": "712613",
            "Monter zabudowy i robót wykończeniowych w budownictwie": "712905",
            "Murarz-tynkarz": "711204",
            "Obuwnik": "753602",
            "Ogrodnik": "611303",
            "Operator maszyn i urządzeń do przetwórstwa tworzyw sztucznych": "814209",
            "Operator maszyn i urządzeń do robót ziemnych i drogowych": "834209",
            "Operator maszyn i urządzeń odlewniczych": "812107",
            "Operator maszyn i urządzeń przemysłu drzewnego (Wprowadzono na podstawie rozporządzenia z dnia 18 stycznia 2023 r.)": "817213",
            "Operator maszyn i urządzeń przemysłu spożywczego": "816003",
            "Operator maszyn i urządzeń przeróbczych": "811205",
            "Operator maszyn i urządzeń w gospodarce odpadami (Wprowadzono na podstawie rozporządzenia z dnia 18 stycznia 2023 r.)": "313211",
            "Operator maszyn leśnych": "834105",
            "Operator maszyn w przemyśle włókienniczym": "815204",
            "Operator obrabiarek skrawających": "722307",
            "Operator procesów introligatorskich": "732305",
            "Operator urządzeń przemysłu ceramicznego": "818115",
            "Operator urządzeń przemysłu chemicznego": "813134",
            "Operator urządzeń przemysłu szklarskiego": "818116",
            "Opiekun medyczny": "532102",
            "Opiekun osoby starszej": "341202",
            "Opiekun w domu pomocy społecznej": "341203",
            "Opiekunka dziecięca": "325905",
            "Opiekunka środowiskowa": "341204",
            "Optyk-mechanik": "731104",
            "Ortoptystka": "325906",
            "Piekarz": "751204",
            "Podolog (opracowano na podstawie dokumentu z dnia 18 czerwca 2021 r.)": "323014",
            "Pracownik obsługi hotelowej": "962907",
            "Pracownik pomocniczy fryzjera": "932920",
            "Pracownik pomocniczy gastronomii": "941203",
            "Pracownik pomocniczy krawca": "932915",
            "Pracownik pomocniczy mechanika": "932916",
            "Pracownik pomocniczy obsługi hotelowej": "911205",
            "Pracownik pomocniczy stolarza": "932918",
            "Pracownik pomocniczy ślusarza": "932917",
            "Pracownik pomocniczy w gospodarce odpadami (Wprowadzono na podstawie rozporządzenia z dnia 18 stycznia 2023 r.)": "932922",
            "Protetyk słuchu": "321401",
            "Przetwórca mięsa": "751108",
            "Przetwórca ryb": "751103",
            "Pszczelarz": "612302",
            "Rękodzielnik wyrobów włókienniczych": "731808",
            "Rolnik": "613003",
            "Rybak śródlądowy": "622201",
            "Sprzedawca": "522301",
            "Stolarz": "752205",
            "Ślusarz": "722204",
            "Tapicer": "753402",
            "": "N/A"
        }
        return zawody_dict.get(specjalnosc, 'N/A')

    def znajdz_indeksy_kolumn(self, arkusz, naglowki):
        # ta funkcja służy do pobrania indeksoów kolumn z arkusza Excela na podstawie nagłówków

        indeksy_kolumn = {}
        for wiersz in range(1, arkusz.max_row + 1):
            for kolumna in range(1, arkusz.max_column + 1):
                komorka = arkusz.cell(row=wiersz, column=kolumna)
                if komorka.value in naglowki:
                    indeksy_kolumn[komorka.value] = kolumna
        return indeksy_kolumn

    def utworz_wykaz(self):
        cwd = os.getcwd()
        parent_dir = os.path.dirname(cwd)
        folder_path = os.path.abspath(os.path.join(
            os.path.dirname(__file__), parent_dir, 'Data', 'Wykazy'))


        wybrane_dane = self.filtruj_dane(self.plik, self.var.get(), self.combobox.get())

        doc = DocxTemplate(os.path.join(cwd, "Szablony", "szablon_wykaz_v3.docx"))


        # naglowki = ['Imię', 'Nazwisko', 'PESEL', 'Data urodzenia', 'Miejsce urodzenia']
        naglowki = ['Imię', 'Nazwisko', 'PESEL']
        
        wb = openpyxl.load_workbook(open(self.plik, "rb"))
        sheet = wb.active
        indeksy_kolumn = self.znajdz_indeksy_kolumn(sheet, naglowki)

        lista = ""

        for linia in range(len(wybrane_dane)):
            rekord = wybrane_dane[linia]

            lista = lista + str(linia + 1) + ". " + rekord[indeksy_kolumn['Imię']-1].value + " " + rekord[indeksy_kolumn['Nazwisko']-1].value + "\n"

        # context = {
        #             'dataWyst': self.data_wystawienia.entry.get(),
        #             'imię': rekord[indeksy_kolumn['Imię']-1].value,
        #             'drugie imię': rekord[indeksy_kolumn['Drugie imię']-1].value,
        #             'nazwisko': rekord[indeksy_kolumn['Nazwisko']-1].value,
        #             # 'dataUrodzenia': rekord[indeksy_kolumn['Data urodzenia']-1].value,
        #             # 'miejsceUrodzenia': rekord[indeksy_kolumn['Miejsce urodzenia']-1].value,
        #             'PESEL': rekord[indeksy_kolumn['PESEL']-1].value,
        #             'zawod': self.combobox.get(),
        #             'kodZawodu': self.symbolZawodu(self.combobox.get()),
        #             'dataRozp': self.data_rozpoczecia.entry.get(),
        #             'dataZako': self.data_zakonczenia.entry.get(),
        #             'godzRozp': self.godzina_rozpoczecia.get()+":"+self.minuty_rozpoczecia.get(),
        #             'stopien': self.var.get(),
        #             'tabela': lista
        #             }

        # Tworzenie kontekstu dla dalszego użycia (np. w szablonach)
        context = {
                    'dataWyst': self.data_wystawienia.entry.get(),
                    'imie': rekord[indeksy_kolumn['Imię']-1].value if 'Imię' in indeksy_kolumn else '',
                    'drugie_imie': rekord[indeksy_kolumn['Drugie imię']-1].value if 'Drugie imię' in indeksy_kolumn else '',
                    'nazwisko': rekord[indeksy_kolumn['Nazwisko']-1].value if 'Nazwisko' in indeksy_kolumn else '',
                    'PESEL': rekord[indeksy_kolumn['PESEL']-1].value if 'PESEL' in indeksy_kolumn else '',
                    'zawod': self.combobox.get(),
                    'kodZawodu': self.symbolZawodu(self.combobox.get()),
                    'dataRozp': self.data_rozpoczecia.entry.get(),
                    'dataZako': self.data_zakonczenia.entry.get(),
                    'godzRozp': f"{self.godzina_rozpoczecia.get()}:{self.minuty_rozpoczecia.get()}",
                    'stopien': self.var.get(),
                    'tabela': lista
        }




        doc.render(context)

        if not os.path.exists(os.path.join(parent_dir, 'Data')):
            os.mkdir(os.path.join(parent_dir, 'Data'))
        if not os.path.exists(os.path.join(parent_dir, 'Data', 'Wykazy')):
            os.mkdir(os.path.join(parent_dir, 'Data', 'Wykazy'))

        doc.save(os.path.join(parent_dir, 'Data', 'Wykazy', f"{context['stopien']}_{context['zawod']}.docx"))

        self.wynik.configure(text=f"utworzono: {str(linia + 1)} pozycji")

        # Ustawienie napisu na przycisku do generowania pdf
        folder_path = os.path.abspath(os.path.join(
            os.path.dirname(__file__), parent_dir, 'Data', 'Wykazy'))

        files_to_convert = os.listdir(folder_path)
        total_files_wykazy = sum(1 for file_name in files_to_convert if os.path.isfile(
            os.path.join(folder_path, file_name)))

        self.btn_utworz_wykaz_pdf.configure(
            text=f"PDF: {str(total_files_wykazy)} plików")

    def utworz_wykaz_pdf(self):
        cwd = os.getcwd()
        parent_dir = os.path.dirname(cwd)
        folder_path = os.path.abspath(os.path.join(
            os.path.dirname(__file__), parent_dir, 'Data', 'Wykazy'))

        convert(folder_path)

    def utworz_skierowania(self):

        cwd = os.getcwd()
        parent_dir = os.path.dirname(cwd)
        folder_path = os.path.abspath(os.path.join(
            os.path.dirname(__file__), parent_dir, 'Data', 'Skierowania'))


        wybrane_dane = self.filtruj_dane(self.plik, self.var.get(), self.combobox.get())

        doc = DocxTemplate(os.path.join(cwd, "Szablony", "szablon_skierowania_v3.docx"))

        # naglowki = ['Imię', 'Nazwisko', 'PESEL', 'Data urodzenia', 'Miejsce urodzenia']
        naglowki = ['Imię', 'Drugie imię', 'Nazwisko', 'PESEL']
        wb = openpyxl.load_workbook(open(self.plik, "rb"))
        sheet = wb.active
        indeksy_kolumn = self.znajdz_indeksy_kolumn(sheet, naglowki)
       
        lista = ""

        for linia in range(len(wybrane_dane)):
            rekord = wybrane_dane[linia]

            context = {
                        'dataWyst': self.data_wystawienia.entry.get(),
                        'imie': rekord[indeksy_kolumn['Imię']-1].value,
                        'drugie_imie': rekord[indeksy_kolumn['Drugie imię']-1].value,
                        'nazwisko': rekord[indeksy_kolumn['Nazwisko']-1].value,
                        # 'dataUrodzenia': rekord[indeksy_kolumn['Data urodzenia']-1].value,
                        # 'miejsceUrodzenia': rekord[indeksy_kolumn['Miejsce urodzenia']-1].value,
                        'PESEL': rekord[indeksy_kolumn['PESEL']-1].value,
                        'zawod': self.combobox.get(),
                        'kodZawodu': self.symbolZawodu(self.combobox.get()),
                        'dataRozp': self.data_rozpoczecia.entry.get(),
                        'dataZako': self.data_zakonczenia.entry.get(),
                        'godzRozp': self.godzina_rozpoczecia.get()+":"+self.minuty_rozpoczecia.get(),
                        'stopien': self.var.get(),
                        'tabela': lista
                        }

            doc.render(context)

            if not os.path.exists(os.path.join(parent_dir, 'Data')):
                os.mkdir(os.path.join(parent_dir, 'Data'))

            if not os.path.exists(os.path.join(parent_dir, 'Data', 'Skierowania')):
                os.mkdir(os.path.join(parent_dir, 'Data', 'Skierowania'))

            doc.save(os.path.join(parent_dir, 'Data', 'Skierowania', f"{context['stopien']}_{context['zawod']}{context['imie']}{context['nazwisko']}.docx"))
            wb.close()
            
        # informacja zwrotna
        self.wynik.configure(
            text=f"utworzono: {str(linia + 1)} dokumentów")

        # Ustawienie napisu na przycisku do generowania pdf
        folder_path_skierowania = os.path.abspath(os.path.join(
            os.path.dirname(__file__), parent_dir, 'Data', 'Skierowania'))

        # Zliczanie plików i folderów w folderze Skierowania


        # Zliczanie tylko plików
        files_to_convert_skierowania = os.listdir(folder_path_skierowania)
        total_files_skierowania = sum(1 for file_name in files_to_convert_skierowania if os.path.isfile(
            os.path.join(folder_path_skierowania, file_name)))

        self.btn_utworz_skierowania_pdf.configure(
            text=f"PDF: {str(total_files_skierowania)} plików")

    def utworz_skierowania_pdf(self):
        """
        Funkcja generuje dokumenty PDF na podstawie dokumentów Word znajdujących się w folderze "Skierowania".

        Nie przyjmuje parametrów.

        Nie zwraca nic.
        """
        current_dir = os.getcwd()
        parent_dir = os.path.dirname(current_dir)

        folder_path = os.path.abspath(os.path.join(
            os.path.dirname(__file__), parent_dir, 'Data', 'Skierowania'))

        # self.show_progress_dialog()  # Call the show_progress_dialog function
        thread = threading.Thread(target=convert, args=(folder_path,))
        thread.start()  # Uruchom tę operację w osobnym wątku
        

    def otworz_folder_wykaz(self):
        """
        Otwiera folder Wykazu w Eksploratorze Windows.

        Ta funkcja używa biblioteki os do uzyskania aktualnego
        katalogu robczego, a następnie używa funkcji os.path.dirname,
        aby znaleźć katalog nadrzędny. Następnie używa funkcji
        os.path.join, aby uzyskać pełną ścieżkę do folderu Wykazu,
        który jest lokalizowany w folderze danych w podkatalogu Wykaz.

        Funkcja używa funkcji subprocess.Popen, aby otworzyć
        Eksplorator Windows i wyświetlić folder Wykazu.

        """
        current_dir = os.getcwd()
        parent_dir = os.path.dirname(current_dir)

        wykazy_folder_path = os.path.abspath(os.path.join(
            os.path.dirname(__file__), parent_dir, 'Data', 'Wykazy'))
        subprocess.Popen(f'explorer "{wykazy_folder_path}"')

    def otworz_folder_skierowania(self):
        """
        Otwiera folder Skierowania w Eksploratorze Windows.

        Ta funkcja używa biblioteki os do uzyskania aktualnego katalogu roboczego,
        a następnie używa funkcji os.path.dirname, aby znaleźć katalog nadrzędny.
        Następnie używa funkcji os.path.join, aby uzyskać pełną ścieżkę do folderu Skierowania,
        który jest lokalizowany w folderze danych w podkatalogu Skierowania.

        Funkcja używa funkcji subprocess.Popen, aby otworzyć Eksplorator Windows
        i wyświetlić folder Skierowania.

        """
        current_dir = os.getcwd()
        parent_dir = os.path.dirname(current_dir)
        folder_path = os.path.abspath(os.path.join(
            os.path.dirname(__file__), parent_dir, 'Data', 'Skierowania'))
        subprocess.Popen(f'explorer "{folder_path}"')
        subprocess.Popen(f'explorer "{folder_path}"')

    def credits(self):
        """Wyświetlanie informacji o autorze i licencji."""
        version_number = self.root.title()
        credits_text = (
            "Autor: Piotr Dębowski\n"
            "Zespół Szkół Energetycznych i Usługowych w Łaziskch Górnych\n\n"
            f"Wersja: {version_number}\n\n")
       
        license_text = "Ten program jest udostępniony na zasadach Licencji MIT. Oznacza to, że jesteś uprawniony do korzystania z niego w ramach warunków określonych w tej licencji.\n\n"
        disclaimer_text = "Autor tego programu nie ponosi odpowiedzialności za ewentualne szkody wynikające z jego użytkowania. Chociaż zrobiłem wszystko, co w mojej mocy, aby zapewnić poprawność i użyteczność tego oprogramowania, nie możgę zagwarantować jego bezbłędnego działania we wszystkich sytuacjach. Korzystając z tego programu, akceptujesz ryzyko związane z jego użyciem.\n\n"
        credits_text += license_text + disclaimer_text
        self.pole_tekstowe.delete("1.0", tk.END)
        self.pole_tekstowe.insert(tk.END, credits_text)



    # def show_progress_dialog(self):
    #     """
    #     Funkcja wyświetla okno modalne z paskiem postępu.

    #     Nie przyjmuje parametrów.

    #     Nie zwraca nic.
    #     """
    #     dialog = tk.Toplevel(self.root)
    #     dialog.grab_set()  # Ustaw okno jako modalne
    #     dialog.resizable(False, False)  # Zablokuj zmianę rozmiaru okna
    #     dialog.title("Generowanie dokumentów PDF")

    #     progress_bar = ttk.Progressbar(dialog, mode='indeterminate')
    #     progress_bar.pack(padx=10, pady=10, fill="x")

    #     cancel_button = ttk.Button(dialog, text="Anuluj", command=dialog.destroy)
    #     cancel_button.pack(padx=10, pady=10)




if __name__ == "__main__":
    app = App()
    app.root.mainloop()
